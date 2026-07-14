from pathlib import Path

from openpyxl import load_workbook

from scripts.create_human_validation_workbooks import (
    LABEL_OPTIONS,
    ValidationRecord,
    build_workbook,
    extract_claim_window,
)


def test_extract_claim_window_uses_target_sentence_and_context():
    text = (
        "বাংলাদেশ ব্যাংক সুদের হার বাড়িয়েছে। "
        "এতে ঋণের খরচ বেড়েছে। "
        "তবে ব্যাংকগুলো ভিন্ন মত দিয়েছে। "
        "প্রতিবেদনটি ২০২৫ সালে প্রকাশিত হয়েছে।"
    )

    focus, context = extract_claim_window(text, 2)

    assert focus == "তবে ব্যাংকগুলো ভিন্ন মত দিয়েছে।"
    assert "Before: বাংলাদেশ ব্যাংক সুদের হার বাড়িয়েছে। এতে ঋণের খরচ বেড়েছে।" in context
    assert "After: প্রতিবেদনটি ২০২৫ সালে প্রকাশিত হয়েছে।" in context


def test_build_workbook_creates_claim_centric_sheets(tmp_path):
    record = ValidationRecord(
        sample_id="hv_001",
        pair_id="pair_001",
        role="original",
        article_id="article_001",
        original_id="article_001",
        perturbation_type="none",
        difficulty="",
        hop_count="",
        newspaper="Test",
        publication_date="2025-01-01",
        headline="বাংলাদেশ ব্যাংক সুদের হার বাড়িয়েছে",
        claim_focus="বাংলাদেশ ব্যাংক সুদের হার বাড়িয়েছে।",
        evidence_context="Before: নীতিগত আলোচনার পর সিদ্ধান্ত নেওয়া হয়।",
        text="বাংলাদেশ ব্যাংক সুদের হার বাড়িয়েছে। নীতিগত আলোচনার পর সিদ্ধান্ত নেওয়া হয়।",
        claim_type="policy_reversal",
    )

    output_path = Path(tmp_path) / "validation.xlsx"
    build_workbook([record], output_path)

    wb = load_workbook(output_path, read_only=False, data_only=True)
    assert wb.sheetnames == ["Instructions", "Samples", "Full Articles"]

    assert LABEL_OPTIONS == ("original", "perturbed", "not sure")

    samples = wb["Samples"]
    sample_headers = [cell.value for cell in next(samples.iter_rows(min_row=1, max_row=1))]
    assert sample_headers == [
        "sample_id",
        "headline",
        "claim_focus",
        "evidence_context",
        "full_article",
        "newspaper",
        "publication_date",
        "annotation_label",
        "annotation_confidence",
        "annotation_justification",
        "claim_type",
    ]

    validation_formulas = [dv.formula1 for dv in samples.data_validations.dataValidation]
    assert any("original,perturbed,not sure" in str(formula) for formula in validation_formulas)

    assert samples["E2"].value == record.text
    assert samples.column_dimensions["K"].hidden is True

    full_articles = wb["Full Articles"]
    full_headers = [cell.value for cell in next(full_articles.iter_rows(min_row=1, max_row=1))]
    assert full_headers == [
        "sample_id",
        "headline",
        "newspaper",
        "publication_date",
        "full_article",
    ]
