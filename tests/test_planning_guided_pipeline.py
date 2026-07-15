from __future__ import annotations

import csv
from pathlib import Path

from src.generation.claim_extraction import HeuristicClaimExtractor, build_claim_extractor
from src.generation.claim_selection import ClaimRanker, ClaimRankingConfig
from src.generation.exporter import HumanValidationWorkbookBuilder
from src.generation.metadata import Article, SampleRecord
from src.generation.models import ModelBundle
from src.generation.perturbation_planner import build_planner
from src.generation.prompts import PLANNING_SCHEMA, build_json_repair_prompt, build_planning_prompt
from src.generation.pipeline import PlanningGuidedRewritePipeline
from src.generation.rewrite_generator import RewriteGenerator
from src.generation.utils import extract_json_payload
from src.generation.verifier import CompositeVerifier, DuplicateVerifier, LocalityVerifier


class FakeGenerator:
    model_name = "fake-instruction-model"
    model_revision = "test"

    def generate_batch(self, prompts, temperatures, seeds, max_new_tokens):
        return ["বাংলাদেশ ব্যাংক নীতিগত সুদের হার ৭ শতাংশ বাড়িয়েছে।" for _ in prompts]


class FakeEmbedder:
    model_name = "fake-sbert"

    def encode(self, texts):
        vectors = []
        for text in texts:
            vectors.append([float(len(text)), float(text.count("বাংলাদেশ")), float(text.count("ব্যাংক")), 1.0])
        return vectors


class FakeNLI:
    model_name = "fake-nli"

    def contradiction_score(self, premise, hypothesis):
        return 0.91 if premise != hypothesis else 0.0


class FakeFluency:
    model_name = "fake-fluency"

    def perplexity(self, text):
        return 35.0


class FakeInstructionModel:
    model_name = "fake-qwen"
    model_revision = "test"

    def __init__(self, response):
        self.response = response

    def generate_text(self, prompt, temperature, seed, max_new_tokens):
        return self.response


class FakeSequenceInstructionModel:
    model_name = "fake-qwen"
    model_revision = "test"

    def __init__(self, responses):
        self.responses = list(responses)

    def generate_text(self, prompt, temperature, seed, max_new_tokens):
        return self.responses.pop(0)


def fake_bundle() -> ModelBundle:
    return ModelBundle(FakeGenerator(), FakeEmbedder(), FakeNLI(), FakeFluency())


def test_claim_extractor_returns_factual_sentence_claims():
    article = Article(
        article_id="a1",
        headline="বাংলাদেশ ব্যাংক সুদের হার বাড়িয়েছে",
        text="বাংলাদেশ ব্যাংক নীতিগত সুদের হার ১০ শতাংশ বাড়িয়েছে। বাজার স্থিতিশীল আছে।",
    )
    claims = HeuristicClaimExtractor().extract(article)

    assert claims
    assert claims[0].sentence_index == 0
    assert claims[0].sentence == "বাংলাদেশ ব্যাংক নীতিগত সুদের হার ১০ শতাংশ বাড়িয়েছে।"
    assert claims[0].numbers == ["১০"]
    assert claims[0].claim_type in {"policy", "numerical"}


def test_ranker_and_planner_create_structured_plan():
    article = Article(
        article_id="a1",
        headline="বাংলাদেশ ব্যাংক সুদের হার বাড়িয়েছে",
        text="বাংলাদেশ ব্যাংক নীতিগত সুদের হার ১০ শতাংশ বাড়িয়েছে।",
    )
    claims = HeuristicClaimExtractor().extract(article)
    ranker = ClaimRanker(ClaimRankingConfig(min_overall_score=0.1, max_risk_score=1.0))
    selected = ranker.select(article, claims)

    assert selected is not None
    assert selected.overall_score > 0
    plan = build_planner().create_plan(selected)
    assert plan.target_claim.sentence_index == 0
    assert plan.edit_scope == "target_sentence"
    assert plan.verification_constraints["preserve_all_other_sentences"] is True


def test_llm_claim_extractor_parses_structured_json():
    article = Article(
        article_id="a1",
        headline="রিজার্ভ বেড়েছে",
        text="বাংলাদেশ ব্যাংকের রিজার্ভ ২৭.০৪ বিলিয়ন ডলারে পৌঁছেছে।",
    )
    model = FakeInstructionModel(
        """
        {
          "claims": [
            {
              "sentence_index": 0,
              "sentence": "বাংলাদেশ ব্যাংকের রিজার্ভ ২৭.০৪ বিলিয়ন ডলারে পৌঁছেছে।",
              "claim": "Bangladesh Bank reserve reached 27.04B USD",
              "type": "numeric",
              "entities": ["বাংলাদেশ ব্যাংক"],
              "numbers": ["২৭.০৪"],
              "policies": [],
              "dates": [],
              "confidence": 0.91
            }
          ]
        }
        """
    )

    claims = build_claim_extractor({"backend": "llm_json", "min_confidence": 0.1}, model=model).extract(article)

    assert len(claims) == 1
    assert claims[0].claim_type == "numerical"
    assert claims[0].claim_text == "Bangladesh Bank reserve reached 27.04B USD"
    assert claims[0].extractor_model == "fake-qwen"


def test_extract_json_payload_finds_embedded_json():
    payload = extract_json_payload('Here is the answer:\n```json\n{"claims": []}\n```\n')

    assert payload == {"claims": []}


def test_llm_claim_extractor_repairs_invalid_json():
    article = Article(
        article_id="a1",
        headline="রিজার্ভ বেড়েছে",
        text="বাংলাদেশ ব্যাংকের রিজার্ভ ২৭.০৪ বিলিয়ন ডলারে পৌঁছেছে।",
    )
    model = FakeSequenceInstructionModel(
        [
            "আমি JSON দিতে পারছি না।",
            """
            {
              "claims": [
                {
                  "sentence_index": 0,
                  "sentence": "বাংলাদেশ ব্যাংকের রিজার্ভ ২৭.০৪ বিলিয়ন ডলারে পৌঁছেছে।",
                  "claim": "Bangladesh Bank reserve reached 27.04B USD",
                  "type": "numeric",
                  "entities": ["বাংলাদেশ ব্যাংক"],
                  "numbers": ["২৭.০৪"],
                  "policies": [],
                  "dates": [],
                  "confidence": 0.91
                }
              ]
            }
            """,
        ]
    )

    claims = build_claim_extractor({"backend": "llm_json", "min_confidence": 0.1}, model=model).extract(article)

    assert len(claims) == 1
    assert claims[0].claim_type == "numerical"


def test_llm_claim_extractor_accepts_single_claim_object():
    article = Article(
        article_id="a1",
        headline="রিজার্ভ বেড়েছে",
        text="বাংলাদেশ ব্যাংকের রিজার্ভ ২৭.০৪ বিলিয়ন ডলারে পৌঁছেছে।",
    )
    model = FakeInstructionModel(
        """
        {
          "sentence_index": 0,
          "sentence": "বাংলাদেশ ব্যাংকের রিজার্ভ ২৭.০৪ বিলিয়ন ডলারে পৌঁছেছে।",
          "claim": "Bangladesh Bank reserve reached 27.04B USD",
          "numbers": ["২৭.০৪"],
          "entities": ["বাংলাদেশ ব্যাংক"]
        }
        """
    )

    claims = build_claim_extractor({"backend": "llm_json", "min_confidence": 0.1}, model=model).extract(article)

    assert len(claims) == 1
    assert claims[0].sentence_index == 0
    assert claims[0].claim_type == "numerical"


def test_llm_claim_extractor_accepts_alternate_claim_list_key():
    article = Article(
        article_id="a1",
        headline="রিজার্ভ বেড়েছে",
        text="বাংলাদেশ ব্যাংকের রিজার্ভ ২৭.০৪ বিলিয়ন ডলারে পৌঁছেছে।",
    )
    model = FakeInstructionModel(
        """
        {
          "factual_claims": [
            {
              "index": 0,
              "sentence": "বাংলাদেশ ব্যাংকের রিজার্ভ ২৭.০৪ বিলিয়ন ডলারে পৌঁছেছে।",
              "claim_type": "numeric",
              "numbers": ["২৭.০৪"]
            }
          ]
        }
        """
    )

    claims = build_claim_extractor({"backend": "llm_json", "min_confidence": 0.1}, model=model).extract(article)

    assert len(claims) == 1
    assert claims[0].sentence_index == 0
    assert claims[0].claim_type == "numerical"


def test_llm_planner_parses_structured_json():
    article = Article(
        article_id="a1",
        headline="বাংলাদেশ ব্যাংক সুদের হার বাড়িয়েছে",
        text="বাংলাদেশ ব্যাংক নীতিগত সুদের হার ১০ শতাংশ বাড়িয়েছে।",
    )
    claims = HeuristicClaimExtractor().extract(article)
    selected = ClaimRanker(ClaimRankingConfig(min_overall_score=0.1, max_risk_score=1.0)).select(article, claims)
    assert selected is not None
    model = FakeInstructionModel(
        """
        {
          "family": "numerical_fact",
          "target_span": "১০ শতাংশ",
          "replacement": "৭ শতাংশ",
          "locality": "target_sentence",
          "edit_instruction": "Change only the selected interest-rate number to ৭ শতাংশ.",
          "expected_change": "The reported interest rate changes from ১০ শতাংশ to ৭ শতাংশ.",
          "verification_constraints": {"preserve_all_other_sentences": true}
        }
        """
    )

    plan = build_planner({"backend": "llm_json", "allowed_families": ["numerical_fact"]}, model=model).create_plan(selected)

    assert plan.family == "numerical_fact"
    assert plan.target_span == "১০ শতাংশ"
    assert plan.replacement == "৭ শতাংশ"
    assert plan.planner_model == "fake-qwen"


def test_llm_planner_normalizes_nonlocal_scope():
    article = Article(
        article_id="a1",
        headline="বাংলাদেশ ব্যাংক সুদের হার বাড়িয়েছে",
        text="বাংলাদেশ ব্যাংক নীতিগত সুদের হার ১০ শতাংশ বাড়িয়েছে।",
    )
    claims = HeuristicClaimExtractor().extract(article)
    selected = ClaimRanker(ClaimRankingConfig(min_overall_score=0.1, max_risk_score=1.0)).select(article, claims)
    assert selected is not None
    model = FakeInstructionModel(
        """
        {
          "family": "numerical_fact",
          "target_span": "১০ শতাংশ",
          "replacement": "৭ শতাংশ",
          "locality": "বাংলাদেশ ব্যাংক নীতিগত সুদের হার ১০ শতাংশ বাড়িয়েছে।",
          "edit_instruction": "Change only the selected interest-rate number to ৭ শতাংশ.",
          "expected_change": "The reported interest rate changes from ১০ শতাংশ to ৭ শতাংশ.",
          "verification_constraints": {"preserve_all_other_sentences": true}
        }
        """
    )

    plan = build_planner({"backend": "llm_json", "allowed_families": ["numerical_fact"]}, model=model).create_plan(selected)

    assert plan.edit_scope == "target_sentence"


def test_planning_prompt_is_compact_and_scope_constrained():
    article = Article(
        article_id="a1",
        headline="বাংলাদেশ ব্যাংক সুদের হার বাড়িয়েছে",
        text="বাংলাদেশ ব্যাংক নীতিগত সুদের হার ১০ শতাংশ বাড়িয়েছে।",
    )
    claims = HeuristicClaimExtractor().extract(article)
    selected = ClaimRanker(ClaimRankingConfig(min_overall_score=0.1, max_risk_score=1.0)).select(article, claims)
    assert selected is not None

    prompt = build_planning_prompt(selected, ["numerical_fact"])

    assert 'locality must be exactly "target_sentence"' in prompt
    assert "Do not put the full claim sentence inside locality." in prompt
    assert "Ranking metadata:" not in prompt


def test_planning_repair_prompt_does_not_use_claim_fallback():
    prompt = build_json_repair_prompt("rewrite planning", PLANNING_SCHEMA, "not json")

    assert '{"claims": []}' not in prompt
    assert "return {}" in prompt


def test_rewrite_generator_splices_only_target_sentence():
    article = Article(
        article_id="a1",
        headline="বাংলাদেশ ব্যাংক সুদের হার বাড়িয়েছে",
        text="বাংলাদেশ ব্যাংক নীতিগত সুদের হার ১০ শতাংশ বাড়িয়েছে। বাজার স্থিতিশীল আছে। বিনিয়োগ বেড়েছে।",
    )
    claims = HeuristicClaimExtractor().extract(article)
    selected = ClaimRanker(ClaimRankingConfig(min_overall_score=0.1, max_risk_score=1.0)).select(article, claims)
    assert selected is not None
    plan = build_planner().create_plan(selected)

    class DriftGenerator:
        model_name = "drifty"
        model_revision = "test"

        def generate_batch(self, prompts, temperatures, seeds, max_new_tokens):
            return [
                "বাংলাদেশ ব্যাংক নীতিগত সুদের হার ৭ শতাংশ বাড়িয়েছে। বাজার অস্থিতিশীল আছে। বিনিয়োগ কমেছে।"
            ]

    rewritten = RewriteGenerator(DriftGenerator()).rewrite(article, plan, temperature=0.0, seed=1, attempt=1)

    assert rewritten.rewritten_article == "বাংলাদেশ ব্যাংক নীতিগত সুদের হার ৭ শতাংশ বাড়িয়েছে। বাজার স্থিতিশীল আছে। বিনিয়োগ বেড়েছে।"
    report = CompositeVerifier([LocalityVerifier()]).verify(article, rewritten.rewritten_article, plan)
    assert report.passed is True


def test_duplicate_verifier_tracks_only_accepted_outputs():
    embedder = FakeEmbedder()
    duplicate = DuplicateVerifier(embedder=embedder, max_similarity=0.99)
    article = Article(article_id="a1", headline="h", text="বাংলাদেশ ব্যাংক ১০ শতাংশ বলেছে।")
    claims = HeuristicClaimExtractor(min_confidence=0.1).extract(article)
    selected = ClaimRanker(ClaimRankingConfig(min_overall_score=0.1, max_risk_score=1.0)).select(article, claims)
    assert selected is not None
    plan = build_planner().create_plan(selected)
    rewritten = "বাংলাদেশ ব্যাংক ৭ শতাংশ বলেছে।"

    first = duplicate.verify(article, rewritten, plan)
    second = duplicate.verify(article, rewritten, plan)
    assert first.passed is True
    assert second.passed is True

    duplicate.accept(rewritten)
    third = duplicate.verify(article, rewritten, plan)
    assert third.passed is False


def test_pipeline_runs_end_to_end_with_injected_models(tmp_path):
    input_csv = tmp_path / "input.csv"
    with open(input_csv, "w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["article_id", "headline", "text"])
        writer.writeheader()
        writer.writerow(
            {
                "article_id": "a1",
                "headline": "বাংলাদেশ ব্যাংক সুদের হার বাড়িয়েছে",
                "text": "বাংলাদেশ ব্যাংক নীতিগত সুদের হার ১০ শতাংশ বাড়িয়েছে। বাজার স্থিতিশীল আছে।",
            }
        )
    config = {
        "paths": {
            "input_csv": str(input_csv),
            "output_dir": str(tmp_path / "out"),
            "checkpoint": str(tmp_path / "out" / "checkpoint.json"),
        },
        "pipeline": {"seed": 7},
        "input": {"id_column": "article_id", "headline_column": "headline", "text_column": "text"},
        "claim_extraction": {"min_confidence": 0.1},
        "claim_ranking": {"min_overall_score": 0.1, "max_risk_score": 1.0},
        "generation": {"temperature": 0.0, "max_new_tokens": 256},
        "regeneration": {"max_attempts": 3, "temperature_step": 0.1},
        "verification": {
            "semantic_similarity_min": 0.1,
            "contradiction_min": 0.5,
            "fluency_max_perplexity": 100.0,
            "duplicate_max_similarity": 0.99,
        },
        "planner": {},
        "human_validation": {"enabled": False},
    }

    override_dir = tmp_path / "override"
    result = PlanningGuidedRewritePipeline(config, model_bundle=fake_bundle()).run(output_dir=str(override_dir))

    assert len(result.samples) == 1
    sample = result.samples[0]
    assert "৭ শতাংশ" in sample.rewritten_article
    assert sample.verification_scores["passed"] is True
    assert (override_dir / "finfact_bd_rewritten.csv").exists()
    assert (override_dir / "metadata.json").exists()
    assert (override_dir / "checkpoint.json").exists()
    assert not (tmp_path / "out" / "checkpoint.json").exists()


def test_human_validation_workbook_is_claim_first(tmp_path):
    sample = SampleRecord(
        sample_id="s1",
        article_id="a1",
        headline="বাংলাদেশ ব্যাংক সুদের হার বাড়িয়েছে",
        original_article="বাংলাদেশ ব্যাংক নীতিগত সুদের হার ১০ শতাংশ বাড়িয়েছে।",
        rewritten_article="বাংলাদেশ ব্যাংক নীতিগত সুদের হার ৭ শতাংশ বাড়িয়েছে।",
        selected_claim={"sentence": "বাংলাদেশ ব্যাংক নীতিগত সুদের হার ১০ শতাংশ বাড়িয়েছে।"},
        claim_index=0,
        claim_type="policy",
        perturbation_family="policy_reversal",
        rewrite_plan={},
        generator_model="fake",
        model_revision="test",
        prompt_version="v",
        temperature=0.0,
        seed=1,
        verification_scores={"passed": True},
        regeneration_attempts=1,
        timestamp="2026-07-14T00:00:00+00:00",
    )
    output = tmp_path / "validation.xlsx"
    HumanValidationWorkbookBuilder(output).build([sample])

    from openpyxl import load_workbook

    wb = load_workbook(output)
    assert wb.sheetnames == ["Instructions", "Samples", "Full Articles"]
    headers = [cell.value for cell in wb["Samples"][1]]
    assert headers == ["sample_id", "headline", "claim_focus", "context_window", "label", "confidence", "justification"]
