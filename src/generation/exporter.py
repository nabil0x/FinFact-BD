from __future__ import annotations

import csv
import json
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List

from src.generation.metadata import SampleRecord
from src.generation.utils import context_window, write_json

logger = logging.getLogger(__name__)


DATASET_FIELDS = [
    "sample_id",
    "article_id",
    "headline",
    "original_article",
    "rewritten_article",
    "selected_claim",
    "claim_index",
    "claim_type",
    "perturbation_family",
    "rewrite_plan",
    "generator_model",
    "model_revision",
    "prompt_version",
    "temperature",
    "seed",
    "verification_scores",
    "regeneration_attempts",
    "timestamp",
]


@dataclass(frozen=True)
class DatasetExporter:
    output_dir: Path

    def export(self, samples: List[SampleRecord], stats: Dict[str, Any]) -> None:
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self._write_csv(self.output_dir / "finfact_bd_rewritten.csv", samples)
        self._write_jsonl(self.output_dir / "finfact_bd_rewritten.jsonl", samples)
        write_json(
            self.output_dir / "metadata.json",
            {
                "pipeline": "controlled_claim_level_llm_rewriting",
                "total_samples": len(samples),
                "stats": stats,
                "schema": DATASET_FIELDS,
            },
        )
        logger.info("Exported %d samples to %s", len(samples), self.output_dir)

    def _write_csv(self, path: Path, samples: List[SampleRecord]) -> None:
        with open(path, "w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=DATASET_FIELDS)
            writer.writeheader()
            for sample in samples:
                writer.writerow(self._flatten(sample.to_dict()))

    def _write_jsonl(self, path: Path, samples: List[SampleRecord]) -> None:
        with open(path, "w", encoding="utf-8") as handle:
            for sample in samples:
                handle.write(json.dumps(sample.to_dict(), ensure_ascii=False) + "\n")

    def _flatten(self, row: Dict[str, Any]) -> Dict[str, Any]:
        flat: Dict[str, Any] = {}
        for key, value in row.items():
            if isinstance(value, (dict, list)):
                flat[key] = json.dumps(value, ensure_ascii=False)
            else:
                flat[key] = value
        return flat


@dataclass(frozen=True)
class HumanValidationWorkbookBuilder:
    output_path: Path
    context_radius: int = 1

    def build(self, samples: List[SampleRecord]) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.worksheet.datavalidation import DataValidation
        except ImportError as exc:
            raise ImportError("Install openpyxl to create human validation workbooks") from exc

        wb = Workbook()
        wb.remove(wb.active)
        instructions = wb.create_sheet("Instructions")
        self._instructions(instructions, Alignment)
        sample_sheet = wb.create_sheet("Samples")
        self._samples(sample_sheet, samples, Alignment, Font, PatternFill, DataValidation)
        full = wb.create_sheet("Full Articles")
        self._full_articles(full, samples, Alignment, Font, PatternFill)
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.output_path)
        logger.info("Wrote human validation workbook to %s", self.output_path)

    def _instructions(self, ws: object, Alignment: object) -> None:
        rows = [
            ("Task", "Judge the claim_focus first. Decide whether it reads original, rewritten, or not sure."),
            ("Read order", "Headline -> claim_focus -> context_window. Use Full Articles only if context is insufficient."),
            ("Labels", "Use original, rewritten, or not sure. Do not search for hidden edits first."),
            ("Confidence", "Choose high, medium, or low and write a one-sentence justification."),
        ]
        ws.append(["Field", "Instruction"])
        for row in rows:
            ws.append(list(row))
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 110
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _samples(self, ws: object, samples: List[SampleRecord], Alignment: object, Font: object, PatternFill: object, DataValidation: object) -> None:
        headers = [
            "sample_id",
            "headline",
            "claim_focus",
            "context_window",
            "label",
            "confidence",
            "justification",
        ]
        ws.append(headers)
        for sample in samples:
            claim = sample.selected_claim
            ws.append(
                [
                    sample.sample_id,
                    sample.headline,
                    claim["sentence"],
                    context_window(sample.rewritten_article, int(sample.claim_index), self.context_radius),
                    "",
                    "",
                    "",
                ]
            )
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        label_validation = DataValidation(type="list", formula1='"original,rewritten,not sure"', allow_blank=False)
        confidence_validation = DataValidation(type="list", formula1='"high,medium,low"', allow_blank=False)
        ws.add_data_validation(label_validation)
        ws.add_data_validation(confidence_validation)
        label_validation.add(f"E2:E{max(ws.max_row, 2)}")
        confidence_validation.add(f"F2:F{max(ws.max_row, 2)}")
        widths = {"A": 16, "B": 42, "C": 70, "D": 80, "E": 14, "F": 14, "G": 48}
        for column, width in widths.items():
            ws.column_dimensions[column].width = width
        ws.freeze_panes = "A2"

    def _full_articles(self, ws: object, samples: List[SampleRecord], Alignment: object, Font: object, PatternFill: object) -> None:
        ws.append(["sample_id", "headline", "rewritten_article", "original_article"])
        for sample in samples:
            ws.append([sample.sample_id, sample.headline, sample.rewritten_article, sample.original_article])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 42
        ws.column_dimensions["C"].width = 110
        ws.column_dimensions["D"].width = 110
