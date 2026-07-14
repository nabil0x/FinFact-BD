from __future__ import annotations

import csv
import json
from pathlib import Path


def inspect_output(output_dir: Path, preview: int = 3, fast: bool = False, skip_workbook: bool = False) -> None:
    if not output_dir.exists():
        raise SystemExit(f"Output directory not found: {output_dir}")
    print(f"Output directory: {output_dir}")
    for path in sorted(output_dir.iterdir()):
        print(f"{path.name}\t{human_bytes(path.stat().st_size)}")
    metadata = read_json(output_dir / "metadata.json")
    checkpoint = read_json(output_dir / "checkpoint.json")
    csv_path = output_dir / "finfact_bd_rewritten.csv"
    if fast:
        rows: list[dict[str, str]] = []
        csv_count = max(0, count_lines(csv_path) - 1)
    else:
        csv_count, rows = read_csv_preview(csv_path, preview)
    jsonl_count = count_lines(output_dir / "finfact_bd_rewritten.jsonl")
    samples = checkpoint.get("samples", []) if checkpoint else []
    failures = checkpoint.get("failures", []) if checkpoint else []
    print(f"metadata_total_samples: {metadata.get('total_samples') if metadata else 'missing'}")
    print(f"metadata_stats: {metadata.get('stats') if metadata else 'missing'}")
    print(f"checkpoint_accepted: {len(samples)}")
    print(f"checkpoint_failures: {len(failures)}")
    print(f"csv_rows: {csv_count}")
    print(f"jsonl_rows: {jsonl_count}")
    for row in rows[:preview]:
        print(
            {
                "sample_id": row.get("sample_id"),
                "article_id": row.get("article_id"),
                "family": row.get("perturbation_family"),
                "attempts": row.get("regeneration_attempts"),
            }
        )
    if failures:
        print("first_failures:")
        for failure in failures[:preview]:
            print(summarize_failure(failure))
    if fast or skip_workbook:
        print("workbook: skipped")
    else:
        inspect_workbook(output_dir / "human_validation.xlsx", preview)


def read_json(path: Path) -> dict:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def read_csv_preview(path: Path, preview: int) -> tuple[int, list[dict[str, str]]]:
    if not path.exists():
        return 0, []
    count = 0
    rows: list[dict[str, str]] = []
    with open(path, "r", encoding="utf-8", newline="") as handle:
        for row in csv.DictReader(handle):
            count += 1
            if len(rows) < preview:
                rows.append(row)
    return count, rows


def count_lines(path: Path) -> int:
    if not path.exists():
        return 0
    with open(path, "r", encoding="utf-8") as handle:
        return sum(1 for _ in handle)


def inspect_workbook(path: Path, preview: int) -> None:
    if not path.exists():
        print("workbook: missing")
        return
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("workbook: present, openpyxl unavailable")
        return
    wb = load_workbook(path, read_only=True)
    print(f"workbook_sheets: {wb.sheetnames}")
    ws = wb["Samples"]
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, preview + 1), values_only=True):
        print({"sample_id": row[0], "headline": row[1], "claim_focus": row[2], "context_window": row[3]})


def summarize_failure(failure: dict[str, object]) -> dict[str, object]:
    summary: dict[str, object] = {
        "article_id": failure.get("article_id"),
        "reason": failure.get("reason"),
    }
    if failure.get("error"):
        summary["error"] = failure.get("error")
    selected = failure.get("selected_claim")
    if isinstance(selected, dict):
        summary["claim_type"] = selected.get("claim_type")
        summary["claim_sentence"] = selected.get("sentence")
    attempts = failure.get("attempts")
    if isinstance(attempts, list):
        summary["attempt_count"] = len(attempts)
        reasons = []
        for attempt in attempts:
            if not isinstance(attempt, dict):
                continue
            verification = attempt.get("verification")
            if isinstance(verification, dict):
                reasons.append(verification.get("reasons"))
            elif attempt.get("error"):
                reasons.append(attempt.get("error"))
        summary["attempt_reasons"] = reasons[:3]
    return summary


def human_bytes(size: int) -> str:
    units = ["B", "KB", "MB", "GB", "TB"]
    value = float(size)
    for unit in units:
        if value < 1024.0:
            return f"{value:.1f}{unit}"
        value /= 1024.0
    return f"{value:.1f}PB"
