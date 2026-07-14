#!/usr/bin/env python3
"""
Create Excel workbooks for FinFact-BD human validation.

The visible annotation artifact is an XLSX workbook because annotators can
work in Excel or Google Sheets, use dropdown labels, and keep free-text
justifications in a structured form. Internal provenance stays in a separate
manifest so annotators never need to touch CSVs.
"""

import argparse
import csv
import json
import random
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).parent.parent))

from src.config import get_output_dir
from src.generation.perturbation_pipeline import (
    DATASET_RELEASE_DATE,
    DATASET_RELEASE_TAG,
    DATASET_VERSION,
    FACT_AWARE_PERTURBATION_TYPES,
)


LABEL_OPTIONS = ("original", "perturbed", "not sure")


@dataclass(frozen=True)
class ValidationRecord:
    sample_id: str
    pair_id: str
    role: str
    article_id: str
    original_id: str
    perturbation_type: str
    difficulty: str
    hop_count: str
    newspaper: str
    publication_date: str
    headline: str
    claim_focus: str
    evidence_context: str
    text: str
    claim_type: str


SENTENCE_SPLIT_RE = re.compile(r"(?<=[।!?])\s+|(?<=\.)\s+|\n+")
CONTEXT_EXPANSION_PREFIXES = (
    "তবে",
    "কিন্তু",
    "যদিও",
    "কারণ",
    "যেহেতু",
    "ফলে",
    "তাই",
    "অতএব",
    "সুতরাং",
    "এদিকে",
    "অন্যদিকে",
    "এছাড়া",
    "এছাড়াও",
    "পরবর্তীতে",
    "এরপর",
    "উল্লেখ্য",
    "অবশেষে",
    "তিনি",
    "তারা",
    "সে",
    "এটি",
    "এটা",
    "এই",
    "সেই",
    "এগুলো",
    "এদের",
    "উনি",
    "এখানে",
    "সেখানে",
    "তখন",
    "তবুও",
)


def load_combined_rows(path: Path) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    with open(path, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows


def split_quota(total: int, parts: int) -> List[int]:
    base = total // parts
    remainder = total % parts
    return [base + (1 if idx < remainder else 0) for idx in range(parts)]


def _is_original(row: Dict[str, str]) -> bool:
    return str(row.get("label", "")).strip() == "0"


def _is_perturbed(row: Dict[str, str]) -> bool:
    return str(row.get("label", "")).strip() == "1"


def parse_sentence_index(value: object) -> Optional[int]:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def normalize_sentence(sentence: str) -> str:
    return re.sub(r"\s+", " ", sentence).strip()


def split_sentences(text: str) -> List[str]:
    stripped = text.strip()
    if not stripped:
        return []

    sentences = [normalize_sentence(part) for part in SENTENCE_SPLIT_RE.split(stripped) if part.strip()]
    return [sentence for sentence in sentences if sentence]


def _should_expand_context(focus_sentence: str) -> bool:
    stripped = focus_sentence.lstrip("“”\"'‘’(【[ ")
    if not stripped:
        return False
    first_token = stripped.split(maxsplit=1)[0]
    first_token = first_token.strip("।!?।,:;()[]{}\"'‘’“”")
    if first_token in CONTEXT_EXPANSION_PREFIXES:
        return True
    if any(stripped.startswith(prefix + " ") or stripped == prefix for prefix in CONTEXT_EXPANSION_PREFIXES):
        return True
    return False


def extract_claim_window(text: str, sentence_index: Optional[int], window: Optional[int] = None) -> Tuple[str, str]:
    sentences = split_sentences(text)
    if not sentences:
        normalized = normalize_sentence(text)
        return normalized, ""

    index = 0 if sentence_index is None else max(0, min(sentence_index, len(sentences) - 1))
    focus_sentence = sentences[index]
    selected_window = 2 if window is None and _should_expand_context(focus_sentence) else (1 if window is None else window)

    before_sentences = sentences[max(0, index - selected_window):index]
    after_sentences = sentences[index + 1:index + 1 + selected_window]
    context_parts: List[str] = []

    if before_sentences:
        context_parts.append("Before: " + " ".join(before_sentences))
    if after_sentences:
        context_parts.append("After: " + " ".join(after_sentences))

    return focus_sentence, "\n".join(context_parts)


def select_validation_pairs(
    rows: List[Dict[str, str]],
    total_pairs: int,
    seed: int,
) -> Tuple[List[Dict[str, str]], Dict[str, object]]:
    """Select matched original/perturbed pairs with unique original ids."""
    rng = random.Random(seed)

    original_lookup = {
        row["article_id"]: row
        for row in rows
        if _is_original(row)
    }
    perturbed_by_original: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    perturbed_by_type: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for row in rows:
        if _is_perturbed(row):
            perturbed_by_original[row["original_id"]].append(row)
            perturbed_by_type[row["perturbation_type"]].append(row)

    quotas = {
        ptype: quota
        for ptype, quota in zip(FACT_AWARE_PERTURBATION_TYPES, split_quota(total_pairs, len(FACT_AWARE_PERTURBATION_TYPES)))
    }

    selected_perturbed: List[Dict[str, str]] = []
    used_original_ids = set()
    shuffled_original_ids = list(perturbed_by_original.keys())
    rng.shuffle(shuffled_original_ids)

    # Greedy pass: walk unique original ids once and always choose the type that
    # still needs the most quota among the options available for that article.
    for original_id in shuffled_original_ids:
        if sum(quotas.values()) == 0:
            break
        candidates = [
            row
            for row in perturbed_by_original[original_id]
            if quotas.get(row["perturbation_type"], 0) > 0
        ]
        if not candidates:
            continue

        max_remaining = max(quotas[row["perturbation_type"]] for row in candidates)
        best = [row for row in candidates if quotas[row["perturbation_type"]] == max_remaining]
        chosen = rng.choice(best)
        selected_perturbed.append(chosen)
        used_original_ids.add(original_id)
        quotas[chosen["perturbation_type"]] -= 1

    # Fallback pass: if any quota remains, search unused originals by family.
    if sum(quotas.values()) > 0:
        for ptype in FACT_AWARE_PERTURBATION_TYPES:
            if quotas[ptype] <= 0:
                continue
            family_candidates = [
                row
                for row in perturbed_by_type[ptype]
                if row["original_id"] not in used_original_ids
            ]
            rng.shuffle(family_candidates)
            while quotas[ptype] > 0 and family_candidates:
                chosen = family_candidates.pop()
                selected_perturbed.append(chosen)
                used_original_ids.add(chosen["original_id"])
                quotas[ptype] -= 1

    if sum(quotas.values()) > 0:
        raise RuntimeError(
            "Unable to construct a unique, balanced human validation sample. "
            f"Remaining quotas: {quotas}"
        )

    records: List[ValidationRecord] = []
    pair_manifest: List[Dict[str, str]] = []

    for pair_index, perturbed_row in enumerate(selected_perturbed, start=1):
        original_row = original_lookup.get(perturbed_row["original_id"])
        if original_row is None:
            raise RuntimeError(f"Missing original row for original_id={perturbed_row['original_id']}")

        focus_sentence_index = parse_sentence_index(perturbed_row.get("sentence_index"))
        pair_id = f"pair_{pair_index:03d}"
        original_sample_id = f"hv_{(pair_index * 2) - 1:03d}"
        perturbed_sample_id = f"hv_{pair_index * 2:03d}"
        claim_type = perturbed_row.get("perturbation_type", "") or "original"

        original_focus, original_context = extract_claim_window(
            original_row.get("text", ""),
            focus_sentence_index,
        )
        perturbed_focus, perturbed_context = extract_claim_window(
            perturbed_row.get("text", ""),
            focus_sentence_index,
        )

        records.append(
            ValidationRecord(
                sample_id=original_sample_id,
                pair_id=pair_id,
                role="original",
                article_id=original_row["article_id"],
                original_id=original_row["original_id"],
                perturbation_type=original_row.get("perturbation_type", "none"),
                difficulty=original_row.get("difficulty", ""),
                hop_count=str(original_row.get("hop_count", "")),
                newspaper=original_row.get("newspaper", ""),
                publication_date=original_row.get("publication_date", ""),
                headline=original_row.get("headline", ""),
                claim_focus=original_focus,
                evidence_context=original_context,
                text=original_row.get("text", ""),
                claim_type=claim_type,
            )
        )
        records.append(
            ValidationRecord(
                sample_id=perturbed_sample_id,
                pair_id=pair_id,
                role="perturbed",
                article_id=perturbed_row["article_id"],
                original_id=perturbed_row["original_id"],
                perturbation_type=perturbed_row.get("perturbation_type", ""),
                difficulty=perturbed_row.get("difficulty", ""),
                hop_count=str(perturbed_row.get("hop_count", "")),
                newspaper=perturbed_row.get("newspaper", ""),
                publication_date=perturbed_row.get("publication_date", ""),
                headline=perturbed_row.get("headline", ""),
                claim_focus=perturbed_focus,
                evidence_context=perturbed_context,
                text=perturbed_row.get("text", ""),
                claim_type=claim_type,
            )
        )

        pair_manifest.append(
            {
                "pair_id": pair_id,
                "original_sample_id": original_sample_id,
                "perturbed_sample_id": perturbed_sample_id,
                "original_article_id": original_row["article_id"],
                "perturbed_article_id": perturbed_row["article_id"],
                "original_id": perturbed_row["original_id"],
                "perturbation_type": perturbed_row.get("perturbation_type", ""),
                "claim_type": claim_type,
                "difficulty": perturbed_row.get("difficulty", ""),
                "hop_count": str(perturbed_row.get("hop_count", "")),
            }
        )

    summary = {
        "total_pairs": total_pairs,
        "total_samples": len(records),
        "original_count": sum(1 for record in records if record.role == "original"),
        "perturbed_count": sum(1 for record in records if record.role == "perturbed"),
        "perturbation_distribution": dict(
            Counter(record.perturbation_type for record in records if record.role == "perturbed")
        ),
        "difficulty_distribution": dict(
            Counter(record.difficulty for record in records if record.role == "perturbed")
        ),
    }

    return records, {"summary": summary, "pairs": pair_manifest}


def make_instruction_rows() -> List[Tuple[str, str]]:
    return [
        (
            "Task",
            "Judge whether the highlighted claim appears original, perturbed, or not sure. Focus on the claim window and nearby context, not on searching the full article for a hidden edit.",
        ),
        (
            "What you see",
            "Each row shows a headline, a claim_focus sentence, nearby evidence_context, the full article text, and an optional Full Articles sheet for deeper review.",
        ),
        (
            "Read order",
            "Start with the headline, then claim_focus, then evidence_context and full_article. Open Full Articles only if you still need more context.",
        ),
        (
            "original",
            "The sample reads like a clean, unmodified news claim. No clear factual edit is visible from the provided claim window and evidence context.",
        ),
        (
            "perturbed",
            "The sample appears to contain a clear factual alteration or other synthetic edit. Examples include wrong number, entity, date, policy direction, comparison, or quotation.",
        ),
        (
            "not sure",
            "You cannot confidently tell whether the sample is original or perturbed from the provided evidence, or the case is borderline / too domain-specific.",
        ),
        (
            "Confidence",
            "Select one confidence level for every row: high, medium, or low.",
        ),
        (
            "Rule 1",
            "Judge each item independently. Do not use external tools, and do not try to infer hidden labels from sample IDs or row order.",
        ),
        (
            "Rule 2",
            "Judge only the highlighted claim. Ignore writing quality, source reputation, politics, and style unless they affect the origin judgment.",
        ),
        (
            "Rule 3",
            "A sample can be fluent and still be perturbed if the claim itself has been edited.",
        ),
        (
            "Rule 4",
            "If the edit is only stylistic, sentiment-level, or a non-factual rewrite, mark original.",
        ),
        (
            "Rule 5",
            "If you are unsure, use not sure rather than guessing.",
        ),
        (
            "Examples",
            "Perturbed patterns include wrong numbers, wrong institutions, reversed comparisons, date shifts, policy flips, and cause-effect inversions. Original patterns include paraphrase, OCR noise, or tone-only changes when no factual edit is visible.",
        ),
        (
            "Output",
            "Fill annotation_label, annotation_confidence, and annotation_justification in the Samples sheet. Keep the justification to 1-2 sentences.",
        ),
    ]


def style_instructions_sheet(ws) -> None:
    ws["A1"] = "FinFact-BD Original vs Perturbed Validation"
    ws["A1"].font = Font(bold=True, size=16, color="1F1F1F")
    ws["A3"] = "How to annotate"
    ws["A3"].font = Font(bold=True, size=12)

    rows = make_instruction_rows()
    start_row = 5
    for idx, (label, text) in enumerate(rows, start=start_row):
        ws.cell(row=idx, column=1, value=label)
        ws.cell(row=idx, column=2, value=text)
        ws.cell(row=idx, column=1).font = Font(bold=True)
        ws.cell(row=idx, column=1).alignment = Alignment(vertical="top")
        ws.cell(row=idx, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    important_row = start_row + len(rows) + 1
    ws[f"A{important_row}"] = "Important"
    ws[f"A{important_row}"].font = Font(bold=True, size=12)
    ws[f"B{important_row}"] = (
        "This workbook is one annotator copy. Save your completed workbook under a new name after annotation."
    )
    ws[f"B{important_row}"].alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 122
    ws.freeze_panes = "A5"
    ws.sheet_view.zoomScale = 95
    ws["A1"].fill = PatternFill("solid", fgColor="D9EAF7")
    ws["A3"].fill = PatternFill("solid", fgColor="EEF5FB")
    ws[f"A{important_row}"].fill = PatternFill("solid", fgColor="FFF2CC")


def write_samples_sheet(ws, records: List[ValidationRecord]) -> None:
    headers = [
        "sample_id",
        "headline",
        "claim_focus",
        "evidence_context",
        "full_article",
        "newspaper",
        "publication_date",
        "annotation_label",
        "annotation_confidence",
        "annotation_justification",
        "claim_type",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(bottom=thin)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for record in records:
        ws.append(
            [
                record.sample_id,
                record.headline,
                record.claim_focus,
                record.evidence_context,
                record.text,
                record.newspaper,
                record.publication_date,
                "",
                "",
                "",
                record.claim_type,
            ]
        )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=11):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=3).fill = PatternFill("solid", fgColor="FFF2CC")
        ws.cell(row=row_idx, column=3).font = Font(bold=True)
        ws.cell(row=row_idx, column=4).fill = PatternFill("solid", fgColor="EAF3FF")
        ws.cell(row=row_idx, column=5).fill = PatternFill("solid", fgColor="F7F7F7")

    # Make the text column readable without manual resizing.
    widths = {
        "A": 12,
        "B": 42,
        "C": 56,
        "D": 58,
        "E": 110,
        "F": 18,
        "G": 16,
        "H": 18,
        "I": 18,
        "J": 42,
        "K": 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{ws.max_row}"
    ws.sheet_view.zoomScale = 90
    ws.row_dimensions[1].height = 24
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 180

    label_validation = DataValidation(
        type="list",
        formula1='"original,perturbed,not sure"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid label",
        error="Choose one of: original, perturbed, not sure.",
    )
    ws.add_data_validation(label_validation)
    label_validation.add(f"H2:H{ws.max_row}")

    confidence_validation = DataValidation(
        type="list",
        formula1='"high,medium,low"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid confidence",
        error="Choose one of: high, medium, low.",
    )
    ws.add_data_validation(confidence_validation)
    confidence_validation.add(f"I2:I{ws.max_row}")

    ws.column_dimensions["K"].hidden = True

    for cell in ws[1]:
        cell.border = border


def write_full_articles_sheet(ws, records: List[ValidationRecord]) -> None:
    headers = [
        "sample_id",
        "headline",
        "newspaper",
        "publication_date",
        "full_article",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(bottom=thin)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for record in records:
        ws.append(
            [
                record.sample_id,
                record.headline,
                record.newspaper,
                record.publication_date,
                record.text,
            ]
        )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "A": 12,
        "B": 42,
        "C": 18,
        "D": 16,
        "E": 110,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{ws.max_row}"
    ws.sheet_view.zoomScale = 90
    ws.row_dimensions[1].height = 24
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 84

    for cell in ws[1]:
        cell.border = border


def build_workbook(records: List[ValidationRecord], output_path: Path) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    instructions = wb.create_sheet("Instructions")
    style_instructions_sheet(instructions)

    samples = wb.create_sheet("Samples")
    write_samples_sheet(samples, records)

    full_articles = wb.create_sheet("Full Articles")
    write_full_articles_sheet(full_articles, records)

    wb.save(output_path)


def write_manifest(manifest: Dict[str, object], output_path: Path) -> None:
    output_path.write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def write_readme(output_dir: Path, annotator_files: List[str]) -> None:
    lines = [
        "# FinFact-BD Human Validation Pack",
        "",
        f"Frozen release: `{DATASET_RELEASE_TAG}` ({DATASET_RELEASE_DATE}).",
        "",
        "Files:",
    ]
    for file_name in annotator_files:
        lines.append(f"- `{file_name}`")
    lines.extend(
        [
            "",
            "Workflow:",
            "1. Open the workbook assigned to the annotator.",
            "2. Start with the `Samples` sheet. Read the headline, claim_focus, evidence_context, and full_article columns first.",
            "3. Use the `full_article` column in the `Samples` sheet first; the `Full Articles` sheet is only for quicker scrolling if you need extra context.",
            "4. Fill `annotation_label`, `annotation_confidence`, and `annotation_justification` for every row.",
            "5. Use `original`, `perturbed`, or `not sure` as the label.",
            "6. Keep `annotation_justification` to 1-2 sentences.",
            "7. Save the completed workbook under a new name.",
            "",
            "The workbook is intentionally xlsx-based so annotators can work comfortably in Excel or Google Sheets.",
        ]
    )
    (output_dir / "README.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Create Excel workbooks for FinFact-BD human validation.")
    parser.add_argument(
        "--input",
        type=Path,
        default=get_output_dir() / "finfact_bd_combined.csv",
        help="Path to finfact_bd_combined.csv",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=get_output_dir() / "human_validation",
        help="Directory where validation workbooks will be written",
    )
    parser.add_argument(
        "--total-pairs",
        type=int,
        default=150,
        help="Number of original/perturbed pairs to include in the annotation set",
    )
    parser.add_argument(
        "--annotators",
        type=int,
        default=3,
        help="Number of annotator workbooks to generate",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=42,
        help="Random seed for pair selection and per-annotator row order",
    )
    args = parser.parse_args()

    rows = load_combined_rows(args.input)
    records, manifest = select_validation_pairs(rows, args.total_pairs, args.seed)

    args.output_dir.mkdir(parents=True, exist_ok=True)

    manifest["input_file"] = str(args.input)
    manifest["output_dir"] = str(args.output_dir)
    manifest["annotators"] = args.annotators
    manifest["seed"] = args.seed
    manifest["dataset_version"] = DATASET_VERSION
    manifest["release_tag"] = DATASET_RELEASE_TAG
    manifest["release_date"] = DATASET_RELEASE_DATE
    manifest["release_state"] = "frozen"
    manifest["validation_protocol"] = {
        "mode": "claim_centric_origin_judgment",
        "sample_sheet": "Samples",
        "optional_sheet": "Full Articles",
        "sample_columns": [
            "sample_id",
            "headline",
            "claim_focus",
            "evidence_context",
            "full_article",
            "newspaper",
            "publication_date",
            "annotation_label",
            "annotation_confidence",
            "annotation_justification",
            "claim_type (hidden)",
        ],
        "label_options": list(LABEL_OPTIONS),
        "confidence_options": ["high", "medium", "low"],
        "confidence_required": True,
        "claim_type_visibility": "hidden",
        "context_window_policy": "adaptive_default_1_expand_to_2_for_discourse_dependent_claims",
        "frozen_release": DATASET_RELEASE_TAG,
    }

    annotator_files: List[str] = []
    for annotator_idx in range(1, args.annotators + 1):
        rng = random.Random(args.seed + annotator_idx * 101)
        shuffled_records = list(records)
        rng.shuffle(shuffled_records)

        file_name = f"finfact_bd_human_validation_annotator_{annotator_idx}.xlsx"
        build_workbook(shuffled_records, args.output_dir / file_name)
        annotator_files.append(file_name)

    manifest["annotator_files"] = annotator_files
    write_manifest(manifest, args.output_dir / "manifest.json")
    write_readme(args.output_dir, annotator_files)

    print(json.dumps(manifest["summary"], indent=2, ensure_ascii=False))
    print(f"Wrote {len(annotator_files)} XLSX workbooks to {args.output_dir}")


if __name__ == "__main__":
    main()
